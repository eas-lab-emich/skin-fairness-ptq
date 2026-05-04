import json
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ── data ──────────────────────────────────────────────────────────────────────

GROUPS = ["FST-1", "FST-2", "FST-3", "FST-4", "FST-5", "FST-6"]
GROUP_KEYS = ["1", "2", "3", "4", "5", "6"]

methods = [
    {
        "name": "FP32 Baseline",
        "top1": 72.689, "wga": 0.6760, "ud": 0.3022, "macro_f1": 0.5781,
        "compression": 1.0,
        "group_acc": {"1":0.6893,"2":0.6760,"3":0.7371,"4":0.7845,"5":0.8421,"6":0.6962},
        "start_bits_wts": None, "start_bits_acts": None,
        "final_bits_wts": "32", "final_bits_acts": "32",
        "iterations": None, "num_evals": None,
    },
    {
        "name": "Uniform W8A8",
        "top1": 71.732, "wga": 0.6709, "ud": 0.2792, "macro_f1": 0.5622,
        "compression": 4.0,
        "group_acc": {"1":0.6893,"2":0.6732,"3":0.7200,"4":0.7724,"5":0.8202,"6":0.6709},
        "start_bits_wts": 8, "start_bits_acts": 8,
        "final_bits_wts": "8 (all)", "final_bits_acts": "8 (all)",
        "iterations": None, "num_evals": None,
    },
    {
        "name": "Uniform W6A6 (start)",
        "top1": 66.278, "wga": 0.5316, "ud": 0.3400, "macro_f1": 0.4062,
        "compression": 5.33,
        "group_acc": {"1":0.6395,"2":0.6397,"3":0.6533,"4":0.7022,"5":0.7763,"6":0.5316},
        "start_bits_wts": 6, "start_bits_acts": 6,
        "final_bits_wts": "6 (all)", "final_bits_acts": "6 (all)",
        "iterations": None, "num_evals": None,
    },
    {
        "name": "Uniform W4A4 (start)",
        "top1": 22.900, "wga": 0.2171, "ud": 0.1439, "macro_f1": 0.0611,
        "compression": 8.0,
        "group_acc": {"1":0.2313,"2":0.2179,"3":0.2171,"4":0.2276,"5":0.2588,"6":0.3165},
        "start_bits_wts": 4, "start_bits_acts": 4,
        "final_bits_wts": "4 (all)", "final_bits_acts": "4 (all)",
        "iterations": None, "num_evals": None,
    },
    {
        "name": "Greedy W4A4 (weights-only budget)",
        "top1": 65.200, "wga": 0.6190, "ud": 0.2555, "macro_f1": 0.0921,
        "compression": 7.99,
        "group_acc": {"1":0.6190,"2":0.6411,"3":0.6305,"4":0.6755,"5":0.7237,"6":0.7468},
        "start_bits_wts": 4, "start_bits_acts": 4,
        "final_bits_wts": "6:1, 5:3, 4:17", "final_bits_acts": "4 (all)",
        "iterations": 11, "num_evals": None,
    },
    {
        "name": "Greedy W4A4 (combined_proxy budget)",
        "top1": 64.320, "wga": 0.6168, "ud": 0.2079, "macro_f1": 0.0893,
        "compression": 7.99,
        "group_acc": {"1":0.6168,"2":0.6383,"3":0.6229,"4":0.6538,"5":0.7105,"6":0.7215},
        "start_bits_wts": 4, "start_bits_acts": 4,
        "final_bits_wts": "6:1, 5:3, 4:17", "final_bits_acts": "6:2, 5:2, 4:17",
        "iterations": 11, "num_evals": None,
    },
    {
        "name": "Greedy W6A6 (combined_proxy budget) ★",
        "top1": 74.979, "wga": 0.7165, "ud": 0.1980, "macro_f1": 0.5727,
        "compression": 4.07,
        "group_acc": {"1":0.7234,"2":0.7165,"3":0.7562,"4":0.7869,"5":0.8289,"6":0.7342},
        "start_bits_wts": 6, "start_bits_acts": 6,
        "final_bits_wts": "8:10, 7:4, 6:7", "final_bits_acts": "8:11, 7:2, 6:8",
        "iterations": 48, "num_evals": 1616,
    },
]

# ── styles ─────────────────────────────────────────────────────────────────────

HEADER_BG   = "1F3864"   # dark navy
SUBHDR_BG   = "2F5597"   # medium blue
BEST_BG     = "E2EFDA"   # light green
BASELINE_BG = "FFF2CC"   # light yellow
UNIFORM_BG  = "F2F2F2"   # light grey
GREEDY_BG   = "DDEEFF"   # light blue

def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thick = Side(style="medium", color="1F3864")
    thin  = Side(style="thin",   color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

PCT  = "0.0%"
PCT2 = "0.00%"
F2   = "0.00"
F4   = "0.0000"

# ── workbook ───────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ══════════════════════════════════════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Summary"

# title row
ws1.merge_cells("A1:N1")
title_cell = ws1["A1"]
title_cell.value = "Mixed-Precision PTQ + Skin-Tone Fairness — Results Summary"
title_cell.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
title_cell.fill  = fill(HEADER_BG)
title_cell.alignment = center
ws1.row_dimensions[1].height = 28

# sub-header row
ws1.merge_cells("A2:N2")
sub = ws1["A2"]
sub.value = "ResNet-18 · Fitz17k (9-class) · Seed 42 · Test Split (n=2,402)"
sub.font  = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
sub.fill  = fill(SUBHDR_BG)
sub.alignment = center

# column headers
headers = [
    "Method", "Start\nWts bits", "Start\nActs bits",
    "Top-1 (%)", "WGA", "UD",
    "Macro-F1", "Weighted-F1",
    "Compression\n(vs FP32 wts)",
    "Iterations", "Evals",
    "Final Wts\ndistribution", "Final Acts\ndistribution",
    "Notes",
]
ws1.row_dimensions[3].height = 36
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col_idx, value=h)
    c.font      = hdr_font(size=10)
    c.fill      = fill(SUBHDR_BG)
    c.alignment = center
    c.border    = thick_bottom()

# method rows
ROW_FILLS = {
    "FP32 Baseline":                               BASELINE_BG,
    "Uniform W8A8":                                UNIFORM_BG,
    "Uniform W6A6 (start)":                        UNIFORM_BG,
    "Uniform W4A4 (start)":                        UNIFORM_BG,
    "Greedy W4A4 (weights-only budget)":           GREEDY_BG,
    "Greedy W4A4 (combined_proxy budget)":         GREEDY_BG,
    "Greedy W6A6 (combined_proxy budget) ★":       BEST_BG,
}

NOTES = {
    "FP32 Baseline":                        "Full-precision reference",
    "Uniform W8A8":                         "Standard PTQ baseline; ~4× compression",
    "Uniform W6A6 (start)":                 "Starting point for W6A6 greedy run",
    "Uniform W4A4 (start)":                 "Starting point for W4A4 greedy runs; near-random accuracy",
    "Greedy W4A4 (weights-only budget)":    "Only weight bits upgraded by search; acts fixed at 4",
    "Greedy W4A4 (combined_proxy budget)":  "Both weights & acts upgraded; best UD among W4A4 runs",
    "Greedy W6A6 (combined_proxy budget) ★":"BEST — surpasses FP32 on top-1 & WGA; macro-F1 preserved; 48 iters",
}

# weighted_f1 values (from finals)
WF1 = {
    "FP32 Baseline":                        0.7363,
    "Uniform W8A8":                         0.7272,
    "Uniform W6A6 (start)":                 0.6581,
    "Uniform W4A4 (start)":                 None,
    "Greedy W4A4 (weights-only budget)":    None,
    "Greedy W4A4 (combined_proxy budget)":  None,
    "Greedy W6A6 (combined_proxy budget) ★":0.7505,
}

for r_idx, m in enumerate(methods, 4):
    bg = ROW_FILLS.get(m["name"], "FFFFFF")
    bold_row = "★" in m["name"]
    row_data = [
        m["name"],
        m["start_bits_wts"] if m["start_bits_wts"] else "—",
        m["start_bits_acts"] if m["start_bits_acts"] else "—",
        m["top1"] / 100,
        m["wga"],
        m["ud"],
        m["macro_f1"],
        WF1.get(m["name"]),
        m["compression"],
        m["iterations"] if m["iterations"] else "—",
        m["num_evals"] if m["num_evals"] else "—",
        m["final_bits_wts"],
        m["final_bits_acts"],
        NOTES.get(m["name"], ""),
    ]
    fmts = [None, None, None, PCT, F4, F4, F4, F4, F2, None, None, None, None, None]

    for c_idx, (val, fmt) in enumerate(zip(row_data, fmts), 1):
        c = ws1.cell(row=r_idx, column=c_idx, value=val)
        c.font      = body_font(bold=bold_row)
        c.fill      = fill(bg)
        c.border    = thin_border()
        c.alignment = center if c_idx > 1 else left
        if fmt and val is not None:
            c.number_format = fmt

ws1.row_dimensions[4].height = 16  # FP32 row

# column widths
col_widths = [38, 9, 9, 10, 10, 10, 10, 11, 13, 10, 8, 18, 18, 42]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# freeze panes
ws1.freeze_panes = "B4"

# conditional formatting — highlight best in Top-1, WGA, UD, Macro-F1 columns
# (just data bars for numeric cols)
data_rows = f"D4:D{3+len(methods)}"
ws1.conditional_formatting.add(
    data_rows,
    DataBarRule(start_type="min", end_type="max",
                color="638EC6", showValue=True)
)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Per-Group Accuracy
# ══════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Per-Group Accuracy")

ws2.merge_cells("A1:I1")
t = ws2["A1"]
t.value = "Per-Group Accuracy by Fitzpatrick Skin Type"
t.font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
t.fill  = fill(HEADER_BG)
t.alignment = center
ws2.row_dimensions[1].height = 26

ws2.merge_cells("A2:I2")
s = ws2["A2"]
s.value = "Fitzpatrick Scale: I/II = lighter · III/IV = medium · V/VI = darker  |  WGA = min across groups  |  UD = max − min"
s.font  = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
s.fill  = fill(SUBHDR_BG)
s.alignment = center

group_headers = ["Method"] + GROUPS + ["WGA (min)", "UD (max−min)"]
ws2.row_dimensions[3].height = 32
for c_idx, h in enumerate(group_headers, 1):
    c = ws2.cell(row=3, column=c_idx, value=h)
    c.font      = hdr_font(size=10)
    c.fill      = fill(SUBHDR_BG)
    c.alignment = center
    c.border    = thick_bottom()

for r_idx, m in enumerate(methods, 4):
    bg      = ROW_FILLS.get(m["name"], "FFFFFF")
    bold_row = "★" in m["name"]
    accs    = [m["group_acc"][k] for k in GROUP_KEYS]
    wga     = min(accs)
    ud      = max(accs) - min(accs)
    row_vals = [m["name"]] + accs + [wga, ud]

    for c_idx, val in enumerate(row_vals, 1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font      = body_font(bold=bold_row)
        c.fill      = fill(bg)
        c.border    = thin_border()
        c.alignment = center if c_idx > 1 else left
        if c_idx > 1:
            c.number_format = PCT

ws2.column_dimensions["A"].width = 38
for i in range(2, len(group_headers) + 1):
    ws2.column_dimensions[get_column_letter(i)].width = 13

# color scale for each group column (green=high)
for col_letter in ["B","C","D","E","F","G","H"]:
    data_range = f"{col_letter}4:{col_letter}{3+len(methods)}"
    ws2.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
    )
# UD column — lower = better, so invert colors
ud_range = f"I4:I{3+len(methods)}"
ws2.conditional_formatting.add(
    ud_range,
    ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B",
    )
)

ws2.freeze_panes = "B4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — W6A6 Greedy Search Trajectory
# ══════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("W6A6 Search Trajectory")

ws3.merge_cells("A1:J1")
t = ws3["A1"]
t.value = "Greedy W6A6 Combined_proxy Budget — Search Trajectory (48 iterations)"
t.font  = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
t.fill  = fill(HEADER_BG)
t.alignment = center
ws3.row_dimensions[1].height = 24

traj_headers = ["Iter", "Layer", "Component", "Bits (old→new)",
                "ΔWGA", "WGA", "UD", "Top-1 (%)", "Macro-F1",
                "Δ Bits used"]
ws3.row_dimensions[2].height = 30
for c_idx, h in enumerate(traj_headers, 1):
    c = ws3.cell(row=2, column=c_idx, value=h)
    c.font      = hdr_font(size=10)
    c.fill      = fill(SUBHDR_BG)
    c.alignment = center
    c.border    = thick_bottom()

with open("/home/mussa/skin-fairness-ptq/search/results/run_greedy_w6a6_combined_budget3_seed42/fairness_search_history.json") as f:
    history = json.load(f)

for r_idx, item in enumerate(history, 3):
    delta_wga = item["delta_wga"]
    row_vals = [
        item["iteration"],
        item["layer"],
        item["component"],
        f"{item['old_bits']}→{item['new_bits']}",
        delta_wga,
        item["wga_after"],
        item["ud_score_after"],
        item["top1_after"] / 100,
        item["macro_f1_after"],
        item["delta_bits"],
    ]
    fmts = [None, None, None, None, "+0.0000;-0.0000", F4, F4, PCT, F4, "#,##0"]
    # row bg: green if positive WGA gain, neutral otherwise
    row_bg = "E2EFDA" if delta_wga > 0.005 else ("FFF2CC" if delta_wga > 0 else "FCE4D6")

    for c_idx, (val, fmt) in enumerate(zip(row_vals, fmts), 1):
        c = ws3.cell(row=r_idx, column=c_idx, value=val)
        c.font      = body_font(size=9)
        c.fill      = fill(row_bg)
        c.border    = thin_border()
        c.alignment = center if c_idx != 2 else left
        if fmt and val is not None:
            c.number_format = fmt

traj_widths = [6, 32, 11, 14, 12, 10, 10, 10, 11, 14]
for i, w in enumerate(traj_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — W4A4 Trajectory (combined_proxy, 11 iters)
# ══════════════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("W4A4 Search Trajectory")

ws4.merge_cells("A1:J1")
t = ws4["A1"]
t.value = "Greedy W4A4 Combined_proxy Budget — Search Trajectory (11 iterations)"
t.font  = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
t.fill  = fill(HEADER_BG)
t.alignment = center
ws4.row_dimensions[1].height = 24

for c_idx, h in enumerate(traj_headers, 1):
    c = ws4.cell(row=2, column=c_idx, value=h)
    c.font      = hdr_font(size=10)
    c.fill      = fill(SUBHDR_BG)
    c.alignment = center
    c.border    = thick_bottom()

with open("/home/mussa/skin-fairness-ptq/search/results/run_greedy_combined_budget4_seed42/fairness_search_history.json") as f:
    history4 = json.load(f)

for r_idx, item in enumerate(history4, 3):
    delta_wga = item["delta_wga"]
    row_vals = [
        item["iteration"],
        item["layer"],
        item["component"],
        f"{item['old_bits']}→{item['new_bits']}",
        delta_wga,
        item["wga_after"],
        item["ud_score_after"],
        item["top1_after"] / 100,
        item["macro_f1_after"],
        item["delta_bits"],
    ]
    fmts = [None, None, None, None, "+0.0000;-0.0000", F4, F4, PCT, F4, "#,##0"]
    row_bg = "E2EFDA" if delta_wga > 0.005 else ("FFF2CC" if delta_wga > 0 else "FCE4D6")

    for c_idx, (val, fmt) in enumerate(zip(row_vals, fmts), 1):
        c = ws4.cell(row=r_idx, column=c_idx, value=val)
        c.font      = body_font(size=9)
        c.fill      = fill(row_bg)
        c.border    = thin_border()
        c.alignment = center if c_idx != 2 else left
        if fmt and val is not None:
            c.number_format = fmt

for i, w in enumerate(traj_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.freeze_panes = "A3"

# ── save ──────────────────────────────────────────────────────────────────────

out_path = "/home/mussa/skin-fairness-ptq/skin_fairness_ptq_results.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
