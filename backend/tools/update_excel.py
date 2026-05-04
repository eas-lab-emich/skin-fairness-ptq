#!/usr/bin/env python3
import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


LEGACY_CUSTOM_RESULTS = {
    "mixed_precision_fakequant": {
        "label": "Legacy custom mixed-precision FakeQuant",
        "original_accuracy": 97.10,
        "final_accuracy": 94.96,
        "accuracy_drop": 2.14,
        "compression_ratio": 5.3281,
    },
    "uniform_fakequant_wts_acts": {
        "label": "Legacy custom uniform FakeQuant (wts+acts)",
        "original_accuracy": 97.10,
        "final_accuracy": 95.29,
        "accuracy_drop": 1.81,
        "compression_ratio": 5.2923,
    },
}


def _json_load(path):
    return json.loads(Path(path).read_text())


def _fmt_list(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return ",".join(str(x) for x in v)
    return str(v)


def _fmt_dist(v):
    if not isinstance(v, dict):
        return ""
    items = sorted(v.items(), key=lambda kv: int(kv[0]), reverse=True)
    return ", ".join(f"{k}:{val}" for k, val in items)


def _to_float(v, default=None):
    try:
        return float(v)
    except Exception:
        return default


def discover_runs(results_root):
    root = Path(results_root)
    runs = []
    for run_dir in sorted(root.glob("run_*"), key=lambda p: p.name, reverse=True):
        final_json = run_dir / "final.json"
        if not final_json.exists():
            continue
        final = _json_load(final_json)
        row = {
            "run_name": run_dir.name,
            "run_dir": str(run_dir),
            "task": final.get("task", "cifar"),
            "search_mode": final.get("search_mode", ""),
            "eval_batches": final.get("eval_batches", ""),
            "baseline_top1": final.get("baseline_top1"),
            "ptq_default_top1": final.get("ptq_default_top1", ""),
            "final_top1": final.get("final_top1"),
            "baseline_macro_f1": final.get("baseline_macro_f1", ""),
            "final_macro_f1": final.get("final_macro_f1", ""),
            "final_weighted_f1": final.get("final_weighted_f1", ""),
            "baseline_unfairness_score": final.get("baseline_unfairness_score", ""),
            "final_unfairness_score": final.get("final_unfairness_score", ""),
            "drop_vs_baseline": final.get("drop_vs_baseline"),
            "epsilon": final.get("budget_epsilon", final.get("epsilon", "")),
            "within_budget": final.get("within_budget", ""),
            "weight_bit_options": final.get("weight_bit_options", final.get("bit_options", "")),
            "activation_bit_options": final.get("activation_bit_options", ""),
            "weights_theoretical_bits": final.get("weights_theoretical_bits", ""),
            "fp32_weights_bits": final.get("fp32_weights_bits", ""),
            "compression_ratio": final.get("compression_ratio_vs_fp32_weights", final.get("compression_ratio_vs_fp32", "")),
            "weights_bit_distribution": final.get("weights_bit_distribution", ""),
            "activations_bit_distribution": final.get("activations_bit_distribution", ""),
            "num_evals_executed": final.get("num_evals_executed", ""),
            "checkpoint": final.get("checkpoint", ""),
            "banding": final.get("banding", ""),
            "raw": final,
        }
        runs.append(row)
    return runs


def discover_uniform_findings(distiller_root):
    results_dir = Path(distiller_root) / "results"
    findings = []
    baseline_path = results_dir / "baseline_metrics.json"
    ptq_path = results_dir / "ptq_metrics.json"
    w4a8_path = results_dir / "ptq_metrics_w4a8.json"
    w8a4_path = results_dir / "ptq_metrics_w8a4.json"
    w4a4_path = results_dir / "ptq_metrics_w4a4.json"

    if baseline_path.exists() and ptq_path.exists():
        b = _json_load(baseline_path)
        p = _json_load(ptq_path)
        drop = _to_float(b.get("top1"), 0.0) - _to_float(p.get("top1"), 0.0)
        findings.append(("Uniform W8/A8", f"Baseline {b.get('top1')} -> PTQ {p.get('top1')} (drop {drop:.3f})"))

    if w4a8_path.exists():
        w4a8 = _json_load(w4a8_path)
        findings.append(("Uniform W4/A8", f"Workable: Top1 {w4a8.get('top1')}, loss {w4a8.get('loss')}"))

    if w8a4_path.exists() and w4a4_path.exists():
        w8a4 = _json_load(w8a4_path)
        w4a4 = _json_load(w4a4_path)
        findings.append(("Uniform A4 sensitivity", f"Collapse observed: W8/A4 Top1 {w8a4.get('top1')}, W4/A4 Top1 {w4a4.get('top1')}"))

    if not findings:
        findings = [
            ("Uniform W8/A8", "Default PTQ is close to baseline (small drop)."),
            ("Uniform W4/A8", "W4/A8 is workable with moderate accuracy drop."),
            ("Uniform A4 sensitivity", "A4 collapses strongly in this setup (W8/A4 and W4/A4)."),
        ]
    return findings


def build_comparison_rows(distiller_root):
    rows = []

    # Legacy custom project rows (hardcoded summary values)
    rows.append({
        "run_name": "legacy_custom_mixed_precision_fakequant",
        "run_dir": "/home/mussa/quant_test/mixed_precision_results.json",
        "task": "legacy_custom",
        "search_mode": "legacy_custom",
        "eval_batches": "",
        "baseline_top1": LEGACY_CUSTOM_RESULTS["mixed_precision_fakequant"]["original_accuracy"],
        "ptq_default_top1": "",
        "final_top1": LEGACY_CUSTOM_RESULTS["mixed_precision_fakequant"]["final_accuracy"],
        "drop_vs_baseline": LEGACY_CUSTOM_RESULTS["mixed_precision_fakequant"]["accuracy_drop"],
        "epsilon": "",
        "within_budget": "",
        "weight_bit_options": "",
        "activation_bit_options": "",
        "weights_theoretical_bits": "",
        "fp32_weights_bits": "",
        "compression_ratio": LEGACY_CUSTOM_RESULTS["mixed_precision_fakequant"]["compression_ratio"],
        "weights_bit_distribution": "",
        "activations_bit_distribution": "",
        "num_evals_executed": "",
        "checkpoint": "legacy_custom",
        "banding": "",
        "raw": {},
    })
    rows.append({
        "run_name": "legacy_custom_uniform_fakequant_wts_acts",
        "run_dir": "/home/mussa/quant_test/quantization_results_fakequant.json",
        "task": "legacy_custom",
        "search_mode": "legacy_custom",
        "eval_batches": "",
        "baseline_top1": LEGACY_CUSTOM_RESULTS["uniform_fakequant_wts_acts"]["original_accuracy"],
        "ptq_default_top1": "",
        "final_top1": LEGACY_CUSTOM_RESULTS["uniform_fakequant_wts_acts"]["final_accuracy"],
        "drop_vs_baseline": LEGACY_CUSTOM_RESULTS["uniform_fakequant_wts_acts"]["accuracy_drop"],
        "epsilon": "",
        "within_budget": "",
        "weight_bit_options": "",
        "activation_bit_options": "",
        "weights_theoretical_bits": "",
        "fp32_weights_bits": "",
        "compression_ratio": LEGACY_CUSTOM_RESULTS["uniform_fakequant_wts_acts"]["compression_ratio"],
        "weights_bit_distribution": "",
        "activations_bit_distribution": "",
        "num_evals_executed": "",
        "checkpoint": "legacy_custom",
        "banding": "",
        "raw": {},
    })

    # Uniform Distiller PTQ rows if available
    results_dir = Path(distiller_root) / "results"
    baseline_path = results_dir / "baseline_metrics.json"
    if baseline_path.exists():
        baseline = _json_load(baseline_path)

        def add_uniform_row(name, metric_file):
            path = results_dir / metric_file
            if not path.exists():
                return
            m = _json_load(path)
            base_top1 = _to_float(baseline.get("top1"), None)
            final_top1 = _to_float(m.get("top1"), None)
            drop = (base_top1 - final_top1) if (base_top1 is not None and final_top1 is not None) else ""
            rows.append({
                "run_name": name,
                "run_dir": str(path),
                "task": "cifar",
                "search_mode": "uniform_ptq",
                "eval_batches": "",
                "baseline_top1": baseline.get("top1", ""),
                "ptq_default_top1": "",
                "final_top1": m.get("top1", ""),
                "drop_vs_baseline": drop,
                "epsilon": "",
                "within_budget": "",
                "weight_bit_options": "",
                "activation_bit_options": "",
                "weights_theoretical_bits": "",
                "fp32_weights_bits": "",
                "compression_ratio": "",
                "weights_bit_distribution": "",
                "activations_bit_distribution": "",
                "num_evals_executed": "",
                "checkpoint": m.get("checkpoint_path_used", m.get("checkpoint_path", "")),
                "banding": "",
                "raw": m,
            })

        add_uniform_row("uniform_ptq_default_w8a8", "ptq_metrics.json")
        add_uniform_row("uniform_ptq_w4a8", "ptq_metrics_w4a8.json")
        add_uniform_row("uniform_ptq_w8a4", "ptq_metrics_w8a4.json")
        add_uniform_row("uniform_ptq_w4a4", "ptq_metrics_w4a4.json")

    return rows


def reset_sheet(wb, title):
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)
    return wb.create_sheet(title)


def style_table(ws, widths):
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_summary_sheet(ws, runs):
    headers = [
        "Run", "Task", "Mode", "Eval batches", "Baseline Top1", "PTQ default Top1", "Final Top1", "Baseline Macro-F1", "Final Macro-F1", "Final Weighted-F1", "Baseline UD", "Final UD", "Drop",
        "Within budget", "Epsilon", "Weight bit opts", "Act bit opts", "Weight bits dist", "Act bits dist",
        "Weight compression ratio vs FP32", "Num evals", "Checkpoint", "Run dir",
    ]
    ws.append(headers)
    for r in runs:
        ws.append([
            r["run_name"],
            r.get("task", "cifar"),
            r["search_mode"],
            r["eval_batches"],
            r["baseline_top1"],
            r["ptq_default_top1"],
            r["final_top1"],
            r.get("baseline_macro_f1", ""),
            r.get("final_macro_f1", ""),
            r.get("final_weighted_f1", ""),
            r.get("baseline_unfairness_score", ""),
            r.get("final_unfairness_score", ""),
            r["drop_vs_baseline"],
            r["within_budget"],
            r["epsilon"],
            _fmt_list(r["weight_bit_options"]),
            _fmt_list(r["activation_bit_options"]),
            _fmt_dist(r["weights_bit_distribution"]),
            _fmt_dist(r["activations_bit_distribution"]),
            r["compression_ratio"],
            r["num_evals_executed"],
            r["checkpoint"],
            r["run_dir"],
        ])
    style_table(
        ws,
        {
            "A": 24, "B": 14, "C": 16, "D": 12, "E": 14, "F": 16, "G": 12, "H": 16, "I": 14,
            "J": 16, "K": 10, "L": 13, "M": 10, "N": 20, "O": 16, "P": 28, "Q": 28, "R": 18,
            "S": 10, "T": 40, "U": 68,
        },
    )


def build_key_findings_sheet(ws, findings, runs):
    ws.append(["Finding", "Details"])
    for title, detail in findings:
        ws.append([title, detail])

    if runs:
        best = max(runs, key=lambda r: _to_float(r.get("final_top1"), -1e9))
        ws.append(["Best run (final Top1)", f"{best['run_name']} -> {best.get('final_top1')}"])

    ws.append(["", ""])
    ws.append(["Legacy Comparison (hardcoded from old custom project)", ""])
    ws.append(["Method", "Original Top1", "Final Top1", "Drop", "Compression ratio"])
    row_start = ws.max_row
    for _, item in LEGACY_CUSTOM_RESULTS.items():
        ws.append([
            item["label"],
            item["original_accuracy"],
            item["final_accuracy"],
            item["accuracy_drop"],
            item["compression_ratio"],
        ])
    for c in ws[row_start + 1]:
        c.font = Font(bold=True)

    ws.append(["Activation sensitivity", "Keep activation search conservative (default 8/7/6, skip early layers)."])
    style_table(ws, {"A": 52, "B": 110, "C": 14, "D": 12, "E": 12})


def build_fairness_sheet(ws, runs):
    fairness_runs = [r for r in runs if r.get("baseline_unfairness_score") != ""]
    headers = [
        "Run", "Task", "Eval batches", "N samples",
        "Baseline Top1", "PTQ W8/A8 Top1", "Final Top1",
        "Baseline UD", "PTQ W8/A8 UD", "Final UD",
        "UD change (baseline→final)", "UD change (baseline→PTQ)",
        "Baseline Macro-F1", "Final Macro-F1",
        "Compression ratio", "Within budget",
    ]
    ws.append(headers)
    for r in fairness_runs:
        raw = r.get("raw", {})
        b_ud = _to_float(r.get("baseline_unfairness_score"))
        f_ud = _to_float(r.get("final_unfairness_score"))
        ptq_ud = _to_float(raw.get("ptq_default_unfairness_score", ""))
        ud_change_final = (f_ud - b_ud) if (f_ud is not None and b_ud is not None) else ""
        ud_change_ptq = (ptq_ud - b_ud) if (ptq_ud is not None and b_ud is not None) else ""
        ws.append([
            r["run_name"],
            r.get("task", ""),
            r["eval_batches"],
            raw.get("baseline_n_samples", ""),
            r["baseline_top1"],
            r["ptq_default_top1"],
            r["final_top1"],
            b_ud,
            ptq_ud if ptq_ud is not None else "",
            f_ud,
            ud_change_final,
            ud_change_ptq,
            r.get("baseline_macro_f1", ""),
            r.get("final_macro_f1", ""),
            r["compression_ratio"],
            r["within_budget"],
        ])
    style_table(ws, {
        "A": 36, "B": 16, "C": 12, "D": 12,
        "E": 14, "F": 16, "G": 12,
        "H": 14, "I": 16, "J": 12,
        "K": 22, "L": 22,
        "M": 18, "N": 16,
        "O": 18, "P": 14,
    })


def build_group_accuracies_sheet(ws, results_root):
    root = Path(results_root)
    FITZ_GROUPS = ["1", "2", "3", "4", "5", "6"]
    headers = ["Run", "Model", "Overall Acc", "UD Score"] + [f"Fitz {g}" for g in FITZ_GROUPS]
    ws.append(headers)

    for run_dir in sorted(root.glob("run_*"), key=lambda p: p.name, reverse=True):
        final_json = run_dir / "final.json"
        if not final_json.exists():
            continue
        final = _json_load(final_json)

        b_groups = final.get("baseline_group_accuracies")
        f_groups = final.get("final_group_accuracies")
        if not b_groups and not f_groups:
            continue

        # baseline row
        if b_groups:
            ws.append(
                [run_dir.name, "Baseline (FP32)",
                 final.get("baseline_accuracy", ""),
                 final.get("baseline_unfairness_score", "")]
                + [b_groups.get(g, "") for g in FITZ_GROUPS]
            )

        # PTQ W8/A8 row — read from ptq_default.json if available
        ptq_json = run_dir / "ptq_default.json"
        if ptq_json.exists():
            ptq = _json_load(ptq_json)
            ptq_groups = ptq.get("group_accuracies")
            if ptq_groups:
                ws.append(
                    [run_dir.name, "PTQ W8/A8",
                     ptq.get("accuracy", ""),
                     ptq.get("unfairness_score", "")]
                    + [ptq_groups.get(g, "") for g in FITZ_GROUPS]
                )

        # final mixed-precision row
        if f_groups:
            ws.append(
                [run_dir.name, "Final (mixed-precision)",
                 final.get("final_accuracy", ""),
                 final.get("final_unfairness_score", "")]
                + [f_groups.get(g, "") for g in FITZ_GROUPS]
            )

        # blank separator between runs
        ws.append([""] * len(headers))

    style_table(ws, {
        "A": 36, "B": 22, "C": 14, "D": 12,
        "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10,
    })


def build_raw_sheet(ws, runs):
    headers = [
        "run_name", "run_dir", "task", "search_mode", "eval_batches", "baseline_top1", "ptq_default_top1", "final_top1",
        "baseline_macro_f1", "final_macro_f1", "final_weighted_f1", "baseline_unfairness_score", "final_unfairness_score", "drop_vs_baseline", "epsilon", "within_budget", "weight_bit_options", "activation_bit_options",
        "weights_theoretical_bits", "fp32_weights_bits", "compression_ratio", "weights_bit_distribution",
        "activations_bit_distribution", "num_evals_executed", "checkpoint", "banding", "raw_json",
    ]
    ws.append(headers)
    for r in runs:
        ws.append([
            r["run_name"], r["run_dir"], r.get("task", "cifar"), r["search_mode"], r["eval_batches"], r["baseline_top1"],
            r["ptq_default_top1"], r["final_top1"], r.get("baseline_macro_f1", ""), r.get("final_macro_f1", ""), r.get("final_weighted_f1", ""),
            r.get("baseline_unfairness_score", ""), r.get("final_unfairness_score", ""),
            r["drop_vs_baseline"], r["epsilon"], r["within_budget"],
            json.dumps(r["weight_bit_options"]), json.dumps(r["activation_bit_options"]), r["weights_theoretical_bits"],
            r["fp32_weights_bits"], r["compression_ratio"], json.dumps(r["weights_bit_distribution"]),
            json.dumps(r["activations_bit_distribution"]), r["num_evals_executed"], r["checkpoint"], r["banding"],
            json.dumps(r["raw"], sort_keys=True),
        ])
    style_table(ws, {"A": 24, "B": 70, "C": 14, "D": 16, "E": 12, "F": 12, "G": 14, "H": 12, "I": 16, "J": 14,
                     "K": 16, "L": 10, "M": 10, "N": 12, "O": 22, "P": 20, "Q": 18, "R": 18, "S": 16, "T": 30,
                     "U": 30, "V": 12, "W": 40, "X": 12, "Y": 110})


def main():
    parser = argparse.ArgumentParser(description="Update Distiller mixed-precision Excel summary")
    parser.add_argument("--results-root", default="/home/mussa/skin-fairness-ptq/backend/results")
    parser.add_argument("--out", default="/home/mussa/skin-fairness-ptq/distiller_mixed_precision_results.xlsx")
    args = parser.parse_args()

    results_root = Path(args.results_root)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    runs = discover_runs(results_root)
    distiller_root = results_root.parent.parent
    findings = discover_uniform_findings(distiller_root)
    comparison_rows = build_comparison_rows(distiller_root)
    summary_rows = runs + comparison_rows

    if out_path.exists():
        wb = load_workbook(out_path)
    else:
        wb = Workbook()

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    ws_summary = reset_sheet(wb, "Summary")
    ws_findings = reset_sheet(wb, "Key Findings")
    ws_fairness = reset_sheet(wb, "Fairness Runs")
    ws_groups = reset_sheet(wb, "Group Accuracies")
    ws_raw = reset_sheet(wb, "Runs (raw)")

    build_summary_sheet(ws_summary, summary_rows)
    build_key_findings_sheet(ws_findings, findings, runs)
    build_fairness_sheet(ws_fairness, runs)
    build_group_accuracies_sheet(ws_groups, results_root)
    build_raw_sheet(ws_raw, runs)

    wb.save(out_path)
    print(f"Processed runs: {len(runs)}")
    print(f"Updated workbook: {out_path}")


if __name__ == "__main__":
    main()
