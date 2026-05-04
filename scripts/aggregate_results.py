"""
aggregate_results.py
====================
Collect all results across models and quantization configs and produce a
master Excel workbook for the paper.

Output sheets:
  1. Summary          — one row per (arch, method, config); top1/WGA/UD/macro_F1/compression
  2. Per-Group        — per-Fitzpatrick-group accuracy for all archs × configs
  3. Greedy Trajectory {arch} — per-iteration greedy history for each arch (one sheet each)

Usage:
  python aggregate_results.py
  python aggregate_results.py --results-root /path/to/results --output results.xlsx

The script auto-discovers result directories under --results-root by scanning for
final.json (greedy runs) and sweep_summary.json / uniform_*.json (uniform sweep runs).
"""

import argparse
import json
from pathlib import Path

import pandas as pd

# ---------------------------------------------------------------------------
FITZPATRICK_GROUPS = ["1", "2", "3", "4", "5", "6"]

# Default paths
DEFAULT_RESULTS_ROOT = "/home/mussa/skin-fairness-ptq/search/results"
DEFAULT_OUTPUT = "/home/mussa/skin-fairness-ptq/skin_fairness_multi_model_results.xlsx"

# Known ResNet18 legacy result directories (absolute paths used in earlier runs)
RESNET18_LEGACY_DIRS = [
    "/home/mussa/skin-fairness-ptq/search/results/run_greedy_budget4_seed42",
    "/home/mussa/skin-fairness-ptq/search/results/run_greedy_combined_budget4_seed42",
    "/home/mussa/skin-fairness-ptq/search/results/run_greedy_w6a6_combined_budget3_seed42",
]


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

def _wga(group_accuracies: dict) -> float:
    if not group_accuracies:
        return float("nan")
    return min(float(v) for v in group_accuracies.values())


def _ud(group_accuracies: dict) -> float:
    if not group_accuracies:
        return float("nan")
    vals = [float(v) for v in group_accuracies.values()]
    return max(vals) - min(vals)


def _load_greedy_final(path: Path) -> dict | None:
    """Load final.json from a greedy run directory."""
    fpath = path / "final.json"
    if not fpath.exists():
        return None
    data = json.loads(fpath.read_text())
    arch = data.get("arch", "resnet18")
    start_w = data.get("start_bits_wts", "?")
    start_a = data.get("start_bits_acts", "?")
    ga = data.get("final_group_accuracies", {})
    baseline_ga = data.get("baseline_group_accuracies", {})
    return {
        "arch": arch,
        "method": f"Greedy W{start_w}A{start_a} ({data.get('budget_mode','?')})",
        "config": f"W{start_w}A{start_a}",
        "type": "greedy",
        "top1": data.get("final_top1", float("nan")),
        "wga": _wga(ga),
        "ud": _ud(ga),
        "macro_f1": data.get("final_macro_f1", float("nan")),
        "compression": data.get("compression_ratio_vs_fp32_weights", float("nan")),
        "iterations": data.get("total_iterations", float("nan")),
        "group_accuracies": ga,
        # Also capture baseline row
        "_baseline": {
            "arch": arch,
            "method": "FP32 Baseline",
            "config": "FP32",
            "type": "baseline",
            "top1": data.get("baseline_top1", float("nan")),
            "wga": _wga(baseline_ga),
            "ud": _ud(baseline_ga),
            "macro_f1": data.get("baseline_macro_f1", float("nan")),
            "compression": 1.0,
            "iterations": float("nan"),
            "group_accuracies": baseline_ga,
        },
        "_run_dir": str(path),
    }


def _load_uniform_sweep(path: Path) -> list[dict]:
    """Load sweep_summary.json or individual uniform_W*A*.json files."""
    rows = []
    summary_path = path / "sweep_summary.json"
    if summary_path.exists():
        data = json.loads(summary_path.read_text())
        arch = data.get("arch", "unknown")
        for tag, cfg in data.get("configs", {}).items():
            rows.append({
                "arch": arch,
                "method": f"Uniform {tag}",
                "config": tag,
                "type": "uniform",
                "top1": cfg.get("top1", float("nan")),
                "wga": cfg.get("wga", float("nan")),
                "ud": cfg.get("ud", float("nan")),
                "macro_f1": cfg.get("macro_f1", float("nan")),
                "compression": float("nan"),
                "iterations": float("nan"),
                "group_accuracies": cfg.get("group_accuracies", {}),
            })
        # Baseline from sweep
        baseline = data.get("baseline", {})
        if baseline:
            ga = baseline.get("group_accuracies", {})
            rows.append({
                "arch": arch,
                "method": "FP32 Baseline",
                "config": "FP32",
                "type": "baseline",
                "top1": baseline.get("top1", float("nan")),
                "wga": _wga(ga),
                "ud": _ud(ga),
                "macro_f1": baseline.get("macro_f1", float("nan")),
                "compression": 1.0,
                "iterations": float("nan"),
                "group_accuracies": ga,
            })
    else:
        # Try individual files
        for fpath in sorted(path.glob("uniform_W*.json")):
            tag = fpath.stem.replace("uniform_", "")
            data = json.loads(fpath.read_text())
            ga = data.get("group_accuracies", {})
            rows.append({
                "arch": data.get("arch", path.parent.name),
                "method": f"Uniform {tag}",
                "config": tag,
                "type": "uniform",
                "top1": data.get("top1", float("nan")),
                "wga": _wga(ga),
                "ud": _ud(ga),
                "macro_f1": data.get("macro_f1", float("nan")),
                "compression": float("nan"),
                "iterations": float("nan"),
                "group_accuracies": ga,
            })
    return rows


def _load_trajectory(path: Path) -> list[dict]:
    """Load greedy search history for trajectory sheet."""
    hpath = path / "fairness_search_history.json"
    if not hpath.exists():
        return []
    history = json.loads(hpath.read_text())
    final = json.loads((path / "final.json").read_text()) if (path / "final.json").exists() else {}
    arch = final.get("arch", "unknown")
    start_w = final.get("start_bits_wts", "?")
    start_a = final.get("start_bits_acts", "?")
    rows = []
    for entry in history:
        row = {
            "arch": arch,
            "run": f"W{start_w}A{start_a}",
            "iteration": entry.get("iteration"),
            "layer": entry.get("layer"),
            "component": entry.get("component"),
            "old_bits": entry.get("old_bits"),
            "new_bits": entry.get("new_bits"),
            "wga_before": entry.get("wga_before"),
            "wga_after": entry.get("wga_after"),
            "delta_wga": entry.get("delta_wga"),
            "compression_ratio": entry.get("compression_ratio_after"),
        }
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

def discover_runs(results_root: Path, extra_dirs: list[str]) -> tuple[list, list, list]:
    """Return (greedy_finals, uniform_rows, trajectory_rows)."""
    greedy_finals = []
    uniform_rows_all = []
    trajectory_rows_all = []

    search_dirs = list(results_root.iterdir()) if results_root.exists() else []
    # Add legacy / extra directories
    for d in extra_dirs:
        p = Path(d)
        if p.exists():
            search_dirs.append(p)

    seen = set()
    for d in search_dirs:
        if not d.is_dir():
            continue
        key = str(d.resolve())
        if key in seen:
            continue
        seen.add(key)

        if (d / "final.json").exists():
            row = _load_greedy_final(d)
            if row:
                greedy_finals.append(row)
                trajectory_rows_all.extend(_load_trajectory(d))

        if (d / "sweep_summary.json").exists() or any(d.glob("uniform_W*.json")):
            uniform_rows_all.extend(_load_uniform_sweep(d))

    return greedy_finals, uniform_rows_all, trajectory_rows_all


# ---------------------------------------------------------------------------
# Build DataFrames
# ---------------------------------------------------------------------------

SUMMARY_COLS = ["arch", "method", "config", "type", "top1", "wga", "ud", "macro_f1", "compression", "iterations"]
PERGROUP_COLS = ["arch", "method", "config"] + [f"FST-{g}" for g in FITZPATRICK_GROUPS]


def build_summary_df(greedy_finals: list, uniform_rows: list) -> pd.DataFrame:
    rows = []
    seen_baselines = set()

    for gr in greedy_finals:
        # Add baseline once per arch
        bl = gr["_baseline"]
        bkey = (bl["arch"], bl["config"])
        if bkey not in seen_baselines and bl["group_accuracies"]:
            rows.append({c: bl.get(c) for c in SUMMARY_COLS})
            seen_baselines.add(bkey)
        rows.append({c: gr.get(c) for c in SUMMARY_COLS})

    for ur in uniform_rows:
        if ur["type"] == "baseline":
            bkey = (ur["arch"], "FP32")
            if bkey not in seen_baselines:
                rows.append({c: ur.get(c) for c in SUMMARY_COLS})
                seen_baselines.add(bkey)
        else:
            rows.append({c: ur.get(c) for c in SUMMARY_COLS})

    df = pd.DataFrame(rows, columns=SUMMARY_COLS)
    df = df.drop_duplicates(subset=["arch", "method", "config"])
    df = df.sort_values(["arch", "type", "config"])
    return df.reset_index(drop=True)


def build_pergroup_df(greedy_finals: list, uniform_rows: list) -> pd.DataFrame:
    rows = []
    seen_baselines = set()

    def _ga_row(d: dict) -> dict:
        ga = d.get("group_accuracies", {})
        row = {"arch": d["arch"], "method": d["method"], "config": d["config"]}
        for g in FITZPATRICK_GROUPS:
            row[f"FST-{g}"] = round(float(ga.get(g, float("nan"))) * 100, 2) if g in ga else float("nan")
        return row

    for gr in greedy_finals:
        bl = gr["_baseline"]
        bkey = (bl["arch"], "FP32")
        if bkey not in seen_baselines and bl["group_accuracies"]:
            rows.append(_ga_row(bl))
            seen_baselines.add(bkey)
        rows.append(_ga_row(gr))

    for ur in uniform_rows:
        if ur["type"] == "baseline":
            bkey = (ur["arch"], "FP32")
            if bkey not in seen_baselines:
                rows.append(_ga_row(ur))
                seen_baselines.add(bkey)
        elif ur["group_accuracies"]:
            rows.append(_ga_row(ur))

    df = pd.DataFrame(rows, columns=PERGROUP_COLS)
    df = df.drop_duplicates(subset=["arch", "method", "config"])
    df = df.sort_values(["arch", "config"])
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(output_path: Path, summary_df: pd.DataFrame, pergroup_df: pd.DataFrame,
                trajectory_rows: list):
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        # Sheet 1: Summary
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Sheet 2: Per-Group Accuracy
        pergroup_df.to_excel(writer, sheet_name="Per-Group Accuracy", index=False)

        # Sheet(s): Trajectory per arch
        if trajectory_rows:
            traj_df = pd.DataFrame(trajectory_rows)
            for arch, grp in traj_df.groupby("arch"):
                sheet_name = f"Traj {arch}"[:31]  # Excel sheet name limit
                grp.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply light formatting
        wb = writer.book
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    print(f"Excel written: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Aggregate all skin fairness results into a master Excel.")
    parser.add_argument("--results-root", default=DEFAULT_RESULTS_ROOT,
                        help="Root directory to scan for result folders.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help="Output Excel file path.")
    parser.add_argument("--extra-dirs", nargs="*", default=RESNET18_LEGACY_DIRS,
                        help="Additional result directories to include (e.g. legacy ResNet18 runs).")
    args = parser.parse_args()

    results_root = Path(args.results_root)
    output_path = Path(args.output)

    print(f"Scanning: {results_root}")
    greedy_finals, uniform_rows, trajectory_rows = discover_runs(results_root, args.extra_dirs or [])

    print(f"Found {len(greedy_finals)} greedy runs, {len(uniform_rows)} uniform rows, "
          f"{len(trajectory_rows)} trajectory entries")

    if not greedy_finals and not uniform_rows:
        print("No results found. Run the pipeline first.")
        return

    summary_df = build_summary_df(greedy_finals, uniform_rows)
    pergroup_df = build_pergroup_df(greedy_finals, uniform_rows)

    print("\n=== Summary ===")
    print(summary_df[["arch", "method", "top1", "wga", "ud", "macro_f1"]].to_string(index=False))

    write_excel(output_path, summary_df, pergroup_df, trajectory_rows)


if __name__ == "__main__":
    main()
